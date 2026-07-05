# -*- coding: utf-8 -*-
"""
importar_ips_edificios.py — Importa edificios y su red (IPs/anexos) desde el
Excel corporativo "Extensiones Actualizadas.xlsx" (hoja ExtensionesIP).

Estructura esperada de la hoja:
    col A: numero de bloque   col B: nombre del edificio (solo en la 1a fila)
    col C: punto/equipo       col D: IP        col E: anexo
    col F: descripcion        col G/H: usuario/clave  (NO se importan)

Comportamiento:
- Crea el edificio si no existe (comparacion sin distinguir mayusculas) y
  reutiliza el existente si ya esta registrado.
- Inserta cada punto en edificio_ips, evitando duplicados exactos
  (mismo edificio + nombre + ip + anexo). Es seguro re-ejecutarlo.
- Normaliza IPs que Excel guardo como numero (192168100101 -> 192.168.100.101).
- Por seguridad NO importa las columnas de usuario/clave: la vista de
  edificios la ve cualquier usuario del sistema.

Uso:
    python importar_ips_edificios.py ["ruta\\al\\Excel.xlsx"]
"""
import os
import re
import sys

from openpyxl import load_workbook

from db import get_db_connection, init_db

RUTA_DEFECTO = os.path.join(
    os.path.expanduser('~'), 'Downloads', 'Extensiones Actualizadas.xlsx'
)
HOJA = 'ExtensionesIP'

# Edificios del Excel que ya existen en la base con otro nombre:
# se reutiliza el registro existente en lugar de crear un duplicado.
ALIAS_EDIFICIOS = {
    'edificio vive20': 'Vive20',
    'edificio magestat': 'Magestad',
    'ecr - cerezos': 'Los Cerezos',
    'esa torre san antonio': 'Torre San Antonio',
}


def celda(v):
    """Convierte una celda a texto limpio ('103.0' -> '103')."""
    if v is None:
        return ''
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def normalizar_ip(v):
    """Repara IPs que Excel convirtio en numero y limpia valores raros."""
    s = celda(v)
    if not s or s == ',':
        return ''
    s = s.replace('\n', ' · ').strip()
    # '192168.172.1' -> '192.168.172.1'
    if re.match(r'^192168\.\d', s):
        s = '192.168.' + s[7:]
    # Digitos corridos: 192168XXXYYY -> 192.168.XXX.YYY
    if re.fullmatch(r'192168\d{4,6}', s):
        resto = s[6:]
        s = f'192.168.{resto[:-3]}.{resto[-3:]}'
    return s


def main():
    ruta = sys.argv[1] if len(sys.argv) > 1 else RUTA_DEFECTO
    if not os.path.exists(ruta):
        print(f'ERROR: no se encontro el Excel: {ruta}')
        return 1

    init_db()  # garantiza que la tabla edificio_ips exista
    wb = load_workbook(ruta, data_only=True, read_only=True)
    if HOJA not in wb.sheetnames:
        print(f'ERROR: el Excel no tiene la hoja "{HOJA}". Hojas: {wb.sheetnames}')
        return 1
    ws = wb[HOJA]

    conn = get_db_connection()
    edificios_nuevos = 0
    edificios_existentes = 0
    puntos_insertados = 0
    puntos_duplicados = 0
    filas_omitidas = 0

    edificio_id = None
    orden = 0

    for fila in ws.iter_rows(min_row=2, values_only=True):
        nombre_edificio = celda(fila[1] if len(fila) > 1 else None)
        nombre_punto = celda(fila[2] if len(fila) > 2 else None)
        ip = normalizar_ip(fila[3] if len(fila) > 3 else None)
        anexo = celda(fila[4] if len(fila) > 4 else None)
        descripcion = celda(fila[5] if len(fila) > 5 else None)

        # ¿Empieza un bloque de edificio nuevo?
        if nombre_edificio:
            nombre_final = re.sub(r'\s{2,}', ' ', nombre_edificio)
            nombre_final = ALIAS_EDIFICIOS.get(nombre_final.lower(), nombre_final)
            row = conn.execute(
                'SELECT id FROM edificios WHERE nombre = ? COLLATE NOCASE',
                (nombre_final,)
            ).fetchone()
            if row:
                edificio_id = row['id']
                edificios_existentes += 1
            else:
                cur = conn.execute(
                    'INSERT INTO edificios (nombre) VALUES (?)', (nombre_final,)
                )
                edificio_id = cur.lastrowid
                edificios_nuevos += 1
            orden = 0

        # Filas antes del primer edificio (anexos de oficina) se omiten.
        if edificio_id is None:
            filas_omitidas += 1
            continue

        # Punto de red: requiere al menos nombre, IP o anexo.
        if not nombre_punto and not ip and not anexo:
            filas_omitidas += 1
            continue
        if not nombre_punto:
            nombre_punto = descripcion or 'Equipo'

        ya_existe = conn.execute('''
            SELECT 1 FROM edificio_ips
            WHERE edificio_id = ? AND nombre = ? COLLATE NOCASE
              AND COALESCE(ip, '') = ? AND COALESCE(anexo, '') = ?
        ''', (edificio_id, nombre_punto, ip, anexo)).fetchone()
        if ya_existe:
            puntos_duplicados += 1
            continue

        orden += 1
        conn.execute('''
            INSERT INTO edificio_ips (edificio_id, nombre, ip, anexo, descripcion, orden)
            VALUES (?, ?, ?, ?, ?, ?)
        ''', (edificio_id, nombre_punto, ip, anexo, descripcion, orden))
        puntos_insertados += 1

    conn.commit()

    total_edificios = conn.execute('SELECT COUNT(*) FROM edificios').fetchone()[0]
    total_puntos = conn.execute('SELECT COUNT(*) FROM edificio_ips').fetchone()[0]
    conn.close()

    print('Importacion completada.')
    print(f'  Edificios nuevos creados:      {edificios_nuevos}')
    print(f'  Edificios ya existentes:       {edificios_existentes}')
    print(f'  Puntos de red insertados:      {puntos_insertados}')
    print(f'  Puntos omitidos (duplicados):  {puntos_duplicados}')
    print(f'  Filas sin datos omitidas:      {filas_omitidas}')
    print(f'  TOTAL en base: {total_edificios} edificios, {total_puntos} puntos de red.')
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
