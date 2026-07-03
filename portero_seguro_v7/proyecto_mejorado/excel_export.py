"""
excel_export.py — Exportación de datos a Excel (.xlsx).

Sigue el mismo patrón que pdf_render.py: las rutas en app.py hacen las
consultas y este módulo solo construye el archivo. Devuelve un BytesIO
listo para enviar con send_file.

Un único helper genérico (construir_libro) recibe título, encabezados y
filas, y produce una hoja con el estilo corporativo (banda naranja,
encabezado destacado, columnas autoajustadas y filtro automático).
"""
import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Paleta corporativa (coherente con el tema de la aplicación)
COLOR_NARANJA = 'E8730C'
COLOR_MARRON = '6B4423'
COLOR_CREMA = 'FDEEDD'


def construir_libro(titulo, encabezados, filas, nombre_hoja='Datos'):
    """Crea un .xlsx con una hoja formateada y devuelve un io.BytesIO.

    titulo:      texto de la banda superior (ej. 'Inventario — Portero Seguro').
    encabezados: lista de nombres de columna.
    filas:       iterable de tuplas/listas, en el mismo orden que encabezados.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = nombre_hoja

    borde = Border(bottom=Side(style='thin', color='DDDDDD'))
    n_cols = len(encabezados)

    # ── Fila 1: banda de título ─────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    celda_titulo = ws.cell(row=1, column=1, value=titulo)
    celda_titulo.font = Font(bold=True, size=14, color='FFFFFF')
    celda_titulo.fill = PatternFill('solid', fgColor=COLOR_NARANJA)
    celda_titulo.alignment = Alignment(vertical='center')
    ws.row_dimensions[1].height = 28

    # ── Fila 2: fecha de generación ─────────────────────────────────────
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    celda_fecha = ws.cell(
        row=2, column=1,
        value=f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    )
    celda_fecha.font = Font(size=9, color=COLOR_MARRON, italic=True)
    celda_fecha.fill = PatternFill('solid', fgColor=COLOR_CREMA)

    # ── Fila 4: encabezados ─────────────────────────────────────────────
    fila_enc = 4
    for col, nombre in enumerate(encabezados, start=1):
        c = ws.cell(row=fila_enc, column=col, value=nombre)
        c.font = Font(bold=True, color='FFFFFF', size=10)
        c.fill = PatternFill('solid', fgColor=COLOR_MARRON)
        c.alignment = Alignment(vertical='center')

    # ── Datos ───────────────────────────────────────────────────────────
    anchos = [len(str(h)) for h in encabezados]
    fila_actual = fila_enc
    for fila in filas:
        fila_actual += 1
        for col, valor in enumerate(fila, start=1):
            c = ws.cell(row=fila_actual, column=col, value=valor)
            c.border = borde
            c.font = Font(size=10)
            largo = len(str(valor)) if valor is not None else 0
            if largo > anchos[col - 1]:
                anchos[col - 1] = largo

    # Autoancho (con tope para no generar columnas kilométricas)
    for i, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(i)].width = min(ancho + 3, 55)

    # Filtro automático sobre encabezados + datos, y panel congelado
    ultima_col = get_column_letter(n_cols)
    ws.auto_filter.ref = f'A{fila_enc}:{ultima_col}{max(fila_actual, fila_enc)}'
    ws.freeze_panes = ws.cell(row=fila_enc + 1, column=1)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
